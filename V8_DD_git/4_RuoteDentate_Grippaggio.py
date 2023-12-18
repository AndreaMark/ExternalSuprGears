import numpy as np
import matplotlib.pyplot as plt
from moduli import ruote_fattorigrippaggio as rfg
from moduli.ruote_interpolazione import interp as ip
from datetime import datetime
from tabulate import tabulate
import openpyxl
import sqlite3
import os
import shutil
import pandas as pd

################################################            VARIABILI GLOBALI           ################################################  

start_time = datetime.now()
out_file = open("Grippaggio.txt", "w")                     # scrivo il file log

######################################################    Importazione Valori dai database
####    IPOTESI DI PROGETTO
####    MATERIALE RUOTE DENTATE
materiale_path = "ipotesipreliminari/materiale.xlsx"
workbook = openpyxl.load_workbook(materiale_path)             
sheet = workbook.active                                     

E_modulus = sheet["C1"].value                               # [N/mm2] modulo elastico
v_poisson = sheet["C2"].value                               # coefficiente di Poisson
HB = sheet["C3"].value                                      # Durezza Brinell
sigma_hlim = sheet["C5"].value                              # [N/mm2] allowable stress number (contact) tab p.128
SH_lim = sheet["C6"].value
sigma_flim = sheet["C7"].value #0.358*HV+231                # [N/mm2] nominal stress number for bending
SF_lim = sheet["C8"].value
rho_val = sheet["C9"].value                                 # [kg/m3]
lambda_val = sheet["C10"].value                             # [N/sK] capacità termica
c_val = sheet["C11"].value                                  # [Nm/kgK] calore specifico




lavorazione_path = "ipotesipreliminari/lavorazione.xlsx"
workbook = openpyxl.load_workbook(lavorazione_path)             
sheet = workbook.active                                     
Ra = sheet["C1"].value





lubrificazione_path = "ipotesipreliminari/lubrificazione.xlsx"
workbook = openpyxl.load_workbook(lubrificazione_path)             
sheet = workbook.active                                     
v_40 = sheet["C1"].value                                    # [mm2/s] viscosità dell'olio lubrificata a 40gradiC
eta_oil = sheet["C4"].value                                 # [mPas] viscosità dinamica 
C = sheet["C5"].value                                       # polialfaolfine/esteri                  





####    VALORI GLOBALI
vg_path = "database/variabili_globali.db"
# Connessione
conn = sqlite3.connect(vg_path)
cur = conn.cursor()
cur.execute("SELECT valori FROM variabili_globali")
vg = [row[0] for row in cur.fetchall()]
theta_p = vg[0]                                         # Angolo di pressione
DentiP = vg[1]
DentiR = vg[2]                                          # Numero di denti di pignone e ruota
tau_re = DentiP/DentiR                                  # Reale rapporto di riduzione
u_real = tau_re**(-1)                                    # Gear ratio   
beta = 0                                                # Angolo d'elica  
conn.close()




####    CARATTERISTICHE GEOMETRICHE NON CORRETTE
cnc_path = "database/caratteristiche_nc.db"
conn = sqlite3.connect(cnc_path)
#   PIGNONE
cur1 = conn.cursor()
cur1.execute("SELECT pignone FROM caratteristiche_nc")
cncP = [row[0] for row in cur1.fetchall()]
#   RUOTA
cur2 = conn.cursor()
cur2.execute("SELECT ruota FROM caratteristiche_nc")
cncR = [row[0] for row in cur2.fetchall()]
####    Pignone
diametro_p_p = 2*cncP[0]        # [mm] diametro primitiva pignone
d_rho_pignone = 2*cncP[1]       # [mm] diametro fondamentale pignone
####    Ruota
diametro_p_r = 2*cncR[0]        # [mm] diametro primitiva ruota
d_rho_ruota = 2*cncR[1]         # [mm] diematro fondamentale ruota
conn.close()




####    POST-CORREZIONE: Valori & Dimensioni
pc_path = "database/post_correzione.db"
conn = sqlite3.connect(pc_path)
#   PIGNONE
cur1 = conn.cursor()
cur1.execute("SELECT pignone FROM post_correzione")
pcP = [row[0] for row in cur1.fetchall()]
#   RUOTA
cur2 = conn.cursor()
cur2.execute("SELECT ruota FROM post_correzione")
pcR = [row[0] for row in cur2.fetchall()]
####    pignone&ruota
raggio_corretto_pignone = pcP[6]*0.001              # [m]
raggio_corretto_ruota = pcR[6]*0.001                # [m]
                
b = pcP[11]                                         # [mm] larghezza di fascia  

Ft = pcP[17]                                    # [N] forza tangenziale/circonferenziale                         

alpha_wt =  pcP[20]                             # [rad] angolo di pressione trasversale di funzionamento
alpha_wt_deg = np.degrees(alpha_wt) 
rho_rel =  pcP[22]                              # [mm] curvatura relativa

alphat1 = pcP[21]                               # [rad] angolo in testa del pignone
alphat2 = pcR[21]                               # [rad] angolo in testa della ruota
#   segmento dei contatti
inizio = pcR[24]
fine = pcP[23]
conn.close()


####    VARIABILI CINEMATICHE
vc_path = "database/variabili_cinematiche.db"
conn = sqlite3.connect(vc_path)
#   PIGNONE
cur1 = conn.cursor()
cur1.execute("SELECT pignone FROM variabili_cinematiche")
vcP = [row[0] for row in cur1.fetchall()]
#   RUOTA
cur2 = conn.cursor()
cur2.execute("SELECT ruota FROM variabili_cinematiche")
vcR = [row[0] for row in cur2.fetchall()]
####    pignone
vel_rot_p = vcP[0]                                  # [1/s] velocità di rotazione pignone
v_p = vel_rot_p*raggio_corretto_pignone             # [m/s] velocità lineare al diametro primitivo/di lavoro pignone
####    ruota
vel_rot_r = vcR[0]                                  # [1/s] velocità di rotazione ruota
v_r = vel_rot_r*raggio_corretto_ruota               # [m/s] velocità lineare al diametro primitivo/di lavoro ruota
conn.close()





####     Coefficienti calcolati
if_path = "database/influence_factorsP.db"
conn = sqlite3.connect(if_path)
cur = conn.cursor()
cur.execute("SELECT valori FROM influence_factorsP")
IF = [row[0] for row in cur.fetchall()]
KA = IF[0]
KV = IF[1]
KH_beta = IF[2]
KH_alpha = IF[4]
KMP =  1
conn.close()










################################################            GRIPPAGGIO V.VULLO GEARS VOL3           ################################################  
####    Grafico e calcolo della velocità di slip
g1 = np.arange(-inizio, 0, 0.01)
vg1 = np.array([np.abs(rfg.vg(v_p, gg, 0,  diametro_p_p)) for gg in g1])
vg_1 = np.max(vg1)

g2 = np.arange(0, fine, 0.01)
vg2 = np.array([rfg.vg(v_r, gg, 0,  diametro_p_r) for gg in g2])
vg_2 = np.max(vg2)

plt.figure(num=1)
plt.plot(g1, vg1, "b", label="pignone")
plt.plot(g2, vg2, "r", label ="ruota")
plt.xlabel('Segmento dei contatti [mm]')
plt.ylabel(r'$v_{g1}, v_{g2}$, slip velocity [m/s]')
plt.grid()
plt.legend()
plt.savefig('slipvelocity.png', dpi=300)

tab_vel = [["pitch line velocity", round(v_p,2)],
           ["slip velocity pignone", round(vg_1,2)],
           ["slip velocity ruota", round(vg_2,2)]]

st0 = "velocità"
st00 = tabulate(tab_vel)
out_file.write(st0 + '\n'), print(st0)
out_file.write(st00 + '\n'), print(st00)





################################################            Coefficienti di contatto          ################################################  
BM1 = np.sqrt(lambda_val*rho_val*c_val*0.001)
BM2 = BM1





################################################            flash temperature theta_fl          ################################################
################################################            Thermoelastic factor & Reduced modulus          ################################################      
Er = (E_modulus)/(1-v_poisson**2)

XM = ((Er**(1/4))/BM1)*np.sqrt(1000)





################################################            Approacch factor          ################################################      
XJ = 1





################################################           Load sharing factor          ################################################         
gamma_A = -u_real*(np.tan(alphat2)/np.tan(alpha_wt) - 1) 
gamma_B = (np.tan(alphat1)/np.tan(alpha_wt)) - 1 - (2*np.pi/(DentiP*np.tan(alpha_wt)))
gamma_D = -u_real*(np.tan(alphat2)/np.tan(alpha_wt)) - 1 + (2*np.pi/(DentiP*np.tan(alpha_wt)))
gamma_E = (np.tan(alphat1)/np.tan(alpha_wt) - 1)

X_gammaAB = (7-2)/15 + (1/3)*(g1-gamma_A)/(gamma_B-gamma_A)
X_gamma = 1
X_gammaDE = (7-2)/15 + (1/3)*(gamma_E-g2)/(gamma_E-gamma_D)





################################################           Geometrical factor          ################################################         
####    Angle factor
x_betaalpha = ip(20, 0.978, 22, 1.007, alpha_wt_deg)               # interpolazione dalla tabella 7.1
# beta_b = 0
# ####    Angle factor
# fatt1 = (np.sin(alpha_wt))**(1/4)
# fatt2 = np.sqrt(np.cos(beta_b))
# fatt3 = fatt1*fatt2
# fatt4 = np.sqrt(np.cos(alpha_wt))
# fatt = fatt3/fatt4
# x_betaalpha = 1.22*fatt               # formula vullo 7.30

XG1 = rfg.calcola_XG(x_betaalpha, u_real, gamma_A, gamma_E)
XG2 = rfg.calcola_XG(x_betaalpha, u_real, gamma_E, gamma_A)





################################################          coefficiente d'attrito          ################################################         
####    transverse unit load
wbt = KA*KV*KH_beta*KH_alpha*KMP*(Ft/b)

####    cumulative velocity vector
vsigmac = 2*v_p*np.sin(alpha_wt)

####    local radius of curvature
rho1 = ((1+gamma_A)/(1 + u_real))*(diametro_p_p/2)*np.sin(alpha_wt)
rho2 = ((u_real-gamma_E)/(1 + u_real))*(diametro_p_r/2)*np.sin(alpha_wt)

####    local relative radius of curvature
rho_rel = (rho1*rho2)/(rho1 + rho2)

####    lubricant factor
XL = C * eta_oil**(-0.05)

####    roughness factor
XR = (Ra)**(0.25)

####    Friction coefficient
MU_m = rfg.calcola_mum(XL, XR, vsigmac, rho_rel, wbt)





################################################          theta_fl          ################################################         
theta_fl1 = rfg.calcola_thetafl(MU_m, XM, XJ, XG1, X_gamma, wbt, v_p, diametro_p_p)
theta_fl2 = rfg.calcola_thetafl(MU_m, XM, XJ, XG2, X_gamma, wbt, v_p, diametro_p_r)

tab_thetafl = [["Coefficienti di contatto", round(BM1,2)],
                ["Thermoelastic factor", round(XM,2)],
                 ["Approacch factor", XJ], 
                 ["Load sharing factor", X_gamma],
                 ["Angle factor", x_betaalpha],
                 ["Posizione adimensionale gammaA", round(gamma_A, 2) ], 
                 ["Posizione adimensionale gammaE", round(gamma_E, 2) ],
                 ["Geometrical factor1", round(XG1,2)],
                 ["Geometrical factor2", round(XG2,2)],
                  ["transverse unit load", round(wbt,2)],
                  ["cumulative velocity vector", round(vsigmac,2)],
                  ["local relative radius of curvature", round(rho_rel,2)],
                  ["lubricant factor", round(XL,2)],
                  ["roughness factor", round(XR,2)],
                  ["Friction coefficient", round(MU_m,2)],
                  ["flash temperature1", round(theta_fl1,2)],
                    ["flash temperature2", round(theta_fl2,2)],
                 ]

st2 = "flash temperature"
st3 = tabulate(tab_thetafl)
out_file.write(st2 + '\n'), print(st2)
out_file.write(st3 + '\n'), print(st3)





################################################          grafico temperatura flash - velocità          ################################################         
vp1 = np.arange(0, v_p, 0.01)
theta_fl1_plot = [rfg.calcola_thetafl(MU_m, XM, XJ, XG1, X_gamma, wbt, vv, diametro_p_p) for vv in vp1]
theta_fl2_plot = [rfg.calcola_thetafl(MU_m, XM, XJ, XG2, X_gamma, wbt, vv, diametro_p_r) for vv in vp1]

plt.figure(num=2)
plt.plot(vp1, theta_fl1_plot , "b", label="pignone")
plt.plot(vp1,theta_fl2_plot , "r", label ="ruota")
plt.xlabel('Pitch line velocity [m/s]')
plt.ylabel(r'$\Theta_{fl}$, flash temperature [C]')
plt.grid()
plt.legend()
plt.savefig('flashtemeprature_v.png', dpi=300)

################################################          grafico temperatura flash - carico          ################################################         
wbt1 = np.arange(0, wbt, 0.01)
theta_fl1_plot1 = [rfg.calcola_thetafl(MU_m, XM, XJ, XG1, X_gamma, ww, v_p, diametro_p_p) for ww in wbt1]
theta_fl2_plot1 = [rfg.calcola_thetafl(MU_m, XM, XJ, XG2, X_gamma, ww, v_p, diametro_p_r) for ww in wbt1]

plt.figure(num=3)
plt.plot(wbt1, theta_fl1_plot1 , "b", label="pignone")
plt.plot(wbt1,theta_fl2_plot1 , "r", label ="ruota")
plt.xlabel('transverse unit load [N/mm]')
plt.ylabel(r'$\Theta_{fl}$, flash temperature [C]')
plt.grid()
plt.legend()
plt.savefig('flashtemeprature_w.png', dpi=300)





################################################          verifica della temperatura flash          ################################################         
####    Banda di pressione hertziana
bh = rfg.calcola_bh(Ft, Er, b, theta_p, rho_rel)

alpha_wn = np.arcsin((np.sin(alpha_wt)*np.cos(beta)))
wbn = wbt/(np.cos(alpha_wn)*np.cos(beta))

####  Numeri di Péclet
Pe1 = rfg.calcola_Pe(vg_1, bh*0.001, rho_val, c_val, lambda_val, np.pi/2)
Pe2 = rfg.calcola_Pe(vg_2, bh*0.001, rho_val, c_val, lambda_val, np.pi/2)

st4 = f"NON è possibile utilizzare la relazione generale 7.20: {round(Pe1,2)}, {round(Pe2,2)}"
st5 = f"È possibile utilizzare la relazione generale 7.20: {round(Pe1,2)}, {round(Pe2,2)}"
if Pe1<5 and Pe2<5:
    out_file.write(st4 + '\n'), print(st4)
else: 
    out_file.write(st5 + '\n'), print(st5)





################################################          interfacial bulk temperature          ################################################         
####    theta_fl_media
theta_fl_media = (theta_fl1 + theta_fl2)/(gamma_E-gamma_A)
st61 = f"La temperatura flash media: {round(theta_fl_media,2)}"
out_file.write(st61 + '\n'), print(st61)

####  lubrication system factor
XS = 0.2                                                    # a bagno d'olio

####  multiple mating pinion factor
XMP = 1

####    overral bulk temeprature
theta_oil = 85                                              # fig 8.4 vullo
theta_M = theta_oil + 0.47*XS*XMP*theta_fl_media

st6 = f"La temperatura all'interfaccia è: {round(theta_M,2)}"
out_file.write(st6 + '\n'), print(st6)

theta_Msp = 0.0021*v_p**2 - 0.1188*v_p + 77.088
st7 = f"Dalla formula sperimentale si ottiene invece: {round(theta_Msp,2)}"
out_file.write(st7 + '\n'), print(st7)





################################################          interfacial contactl temperature          ################################################         
theta_fl = max(theta_fl1, theta_fl2)

theta_B = theta_M + theta_fl

st8 = f"L'olio lubrificante deve resistere alla massima temperatura possibile di: {round(theta_B,2)}"
out_file.write(st8 + '\n'), print(st8)


################################################          RUMORE          ################################################ 
L = rfg.calcola_L(u_real, beta, KV, 1.3, 15)
st9 = f"Il rumore prodotto dal riduttore alla distanza di 1m è pari a: {L} dB"
out_file.write(st9 + '\n'), print(st9)

beta_val = np.arange(0, np.radians(25), 0.01)
l_plot = [rfg.calcola_L(u_real, bb, KV, 1.3, 15) for bb in beta_val]

plt.figure(num=4)
plt.plot(beta_val, l_plot)
plt.plot(beta, L, 'or')
plt.xlabel("angolo d'elica [rad]")
plt.ylabel('livello di rumorosità [dB]')
plt.grid()
plt.savefig('rumorosità.png', dpi=300)





################################################            DATABASE        ################################################
out_file.close() 






####    Grip factors
grip_factors = {
    'fattori':["Coefficienti di contatto",
               "Thermoelastic factor",
               "Approacch factor", 
               "Load sharing factor", 
               "Angle factor", 
               "Posizione adimensionale gammaA",
               "Posizione adimensionale gammaE", 
               "Geometrical factor1",
                "Geometrical factor2",
                "transverse unit load", 
                "cumulative velocity vector", 
                "local relative radius of curvature",
                "lubricant factor", 
                "roughness factor", 
                "Friction coefficient", 
                "flash temperature1", 
                "flash temperature2", ],

    'valori':[   round(BM1,2),
                 round(XM,2),
                 XJ,
                 X_gamma,
                 x_betaalpha,
                 round(gamma_A, 2),
                 round(gamma_E, 2),
                 round(XG1,2),
                 round(XG2,2),
                 round(wbt,2),
                 round(vsigmac,2),
                 round(rho_rel,2),
                 round(XL,2),
                 round(XR,2),
                 round(MU_m,2),
                 round(theta_fl1,2),
                 round(theta_fl2,2)
                 ]
}
index_labels1 = ['r1', 'r2', 'r3', 'r4', 'r5', 
                 'r6', 'r7', 'r8', 'r9', 'r10',
                 'r11', 'r12', 'r13', 'r14', 'r15',
                 'r16', 'r17']

GF = pd.DataFrame(grip_factors, index=index_labels1)

# Tabella su excel sempre facilmente consultabile
GF.to_excel('grip_factorsP.xlsx', index=True)






################################################            ORGANIZZAZIONE FILE        ################################################
# Ottieni il percorso della directory in cui si trova lo script
percorso_script = os.path.dirname(os.path.abspath(__file__))

# Costruisci il percorso completo della cartella di origine
cartella_origine = os.path.join(percorso_script)

# Crea una cartella per le figure, tabelle e log all'interno della directory dello script (senza errori se esistono già)
os.makedirs(os.path.join(cartella_origine, 'figure'), exist_ok=True)
os.makedirs(os.path.join(cartella_origine, 'tabelle'), exist_ok=True)
os.makedirs(os.path.join(cartella_origine, 'log'), exist_ok=True)

# Elenca tutti i file nella cartella di origine
files = os.listdir(cartella_origine)

# Sposta i file nelle cartelle corrispondenti
for file in files:
    if file.endswith('.png') or file.endswith('.jpg'):
        shutil.move(os.path.join(cartella_origine, file), os.path.join(cartella_origine, 'figure', file))
    elif file.endswith('.xlsx'):
        shutil.move(os.path.join(cartella_origine, file), os.path.join(cartella_origine, 'tabelle', file))
    elif file.endswith('.txt'):
        shutil.move(os.path.join(cartella_origine, file), os.path.join(cartella_origine, 'log', file))



end_time = datetime.now()
print('Elapsed Time: {}'.format(end_time - start_time))
plt.show()
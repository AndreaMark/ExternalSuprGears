import numpy as np
from moduli import ruote_fattoriinfluenza as rfi
from moduli import ruote_fattoripitting as rfp
from moduli import ruote_fattoriflessione as rff
from datetime import datetime
from tabulate import tabulate
import openpyxl
import sqlite3
import os
import shutil
import pandas as pd

################################################            VARIABILI GLOBALI           ################################################  

start_time = datetime.now()
out_file = open("ResisenzaR.txt", "w")                     # scrivo il file log





######################################################    Importazione Valori dai database
####    IPOTESI DI PROGETTO
####    MATERIALE RUOTE DENTATE
materiale_path = "ipotesipreliminari/materiale.xlsx"
workbook = openpyxl.load_workbook(materiale_path)             
sheet = workbook.active                                     

E_modulus = sheet["C1"].value                               # [N/mm2] modulo elastico
v_poisson = sheet["C2"].value                               # coefficiente di Poisson
HB = sheet["C3"].value                                      # Durezza Brinell
HV = sheet["C4"].value                                      # Durezza Vickers
#  COEFFICIENTI VULLO TAB 3.2 P. 217
#   pitting
Ah = sheet["C14"].value
Bh = sheet["C15"].value
# bending
Af = sheet["C16"].value
Bf = sheet["C17"].value
# Tensioni ammissibili e fattori di sicurezza
sigma_hlim = HV*Ah + Bh                              # [N/mm2] allowable stress number (contact) tab p.128
SH_lim = sheet["C6"].value
sigma_flim= HV*Af + Bf                             # [N/mm2] nominal stress number for bending
SF_lim = sheet["C8"].value
density = sheet["C9"].value*1e-9                            # [kg/mm3]





lavorazione_path = "ipotesipreliminari/lavorazione.xlsx"
workbook = openpyxl.load_workbook(lavorazione_path)             
sheet = workbook.active                                     
Ra = sheet["C1"].value
Rz = sheet["C2"].value                                      # [mm] Rugosità relativa 0.8 micron
f_bp = sheet["C3"].value                                    # [mm] deviazione trasversale della base del dente, limie dato da v>10m/s
F_beta_x = sheet["C4"].value                                # [mm] diseallinamento massimo iniziale prima del rodaggio





lubrificazione_path = "ipotesipreliminari/lubrificazione.xlsx"
workbook = openpyxl.load_workbook(lubrificazione_path)             
sheet = workbook.active                                     
v_40 = sheet["C1"].value                                    # [mm2/s] viscosità dell'olio lubrificata a 40gradiC





####    VALORI GLOBALI
vg_path = "database/variabili_globali.db"
# Connessione
conn = sqlite3.connect(vg_path)
cur = conn.cursor()
cur.execute("SELECT valori FROM variabili_globali")
vg = [row[0] for row in cur.fetchall()]
theta_p = vg[0]                                         # Angolo di pressione
DentiP = vg[1]
DentiR = vg[2]                                          # Numero di denti di pignone e ruota
tau_re = DentiP/DentiR                                  # Reale rapporto di riduzione
u_real = tau_re**(-1)                                    # Gear ratio   
m_uni = vg[4]                                           # modulo e addendum
a = m_uni
beta = 0                                                # Angolo d'elica  
conn.close()





####    CARATTERISTICHE GEOMETRICHE NON CORRETTE
cnc_path = "database/caratteristiche_nc.db"
conn = sqlite3.connect(cnc_path)
#   PIGNONE
cur1 = conn.cursor()
cur1.execute("SELECT pignone FROM caratteristiche_nc")
cncP = [row[0] for row in cur1.fetchall()]
#   RUOTA
cur2 = conn.cursor()
cur2.execute("SELECT ruota FROM caratteristiche_nc")
cncR = [row[0] for row in cur2.fetchall()]
####    Pignone
diametro_p_p = 2*cncP[0]        # [mm] diametro primitiva pignone
d_rho_pignone = 2*cncP[1]       # [mm] diametro fondamentale pignone
####    Ruota
diametro_p_r = 2*cncR[0]        # [mm] diametro primitiva ruota
d_rho_ruota = 2*cncR[1]         # [mm] diematro fondamentale ruota
conn.close()





####    VARIABILI CINEMATICHE
vc_path = "database/variabili_cinematiche.db"
conn = sqlite3.connect(vc_path)
#   PIGNONE
cur1 = conn.cursor()
cur1.execute("SELECT pignone FROM variabili_cinematiche")
vcP = [row[0] for row in cur1.fetchall()]
#   RUOTA
cur2 = conn.cursor()
cur2.execute("SELECT ruota FROM variabili_cinematiche")
vcR = [row[0] for row in cur2.fetchall()]

vel_rot_s = vcP[0]                              # [1/s] velocità di rotazione
vel_rot_rpm = vel_rot_s * 60 / (2 * np.pi)      # [rpm] velocità di rotazione
Momento_pignone = vcP[1]                        # Nm
Momento_ruota = vcR[1]                          # Nm
conn.close()










####    POST-CORREZIONE: Valori & Dimensioni
pc_path = "database/post_correzione.db"
conn = sqlite3.connect(pc_path)
#   PIGNONE
cur1 = conn.cursor()
cur1.execute("SELECT pignone FROM post_correzione")
pcP = [row[0] for row in cur1.fetchall()]
#   RUOTA
cur2 = conn.cursor()
cur2.execute("SELECT ruota FROM post_correzione")
pcR = [row[0] for row in cur2.fetchall()]
####    pignone&ruota
x_p, x_r = pcP[0], pcR[0]                           # fattori correttivi di ruota e pignone
theta_l = pcP[1]                                    # [rad] angolo di lavoro                                                                 
m_corr = pcP[3]                                     # [mm] modulo corretto                                                
ded_p, ded_r = pcP[5], pcR[5]                       # [mm] dedendum pignone e ruota
raggio_corretto_pignone = pcP[6]*0.001              # [m]
raggio_corretto_ruota = pcR[6]*0.001                # [m]
diametro_l_p = 2*pcP[6]                             # [mm] diametro primitiva di lavoro pignone
diametro_l_r = 2*pcR[6]                             # [mm] diametro primitiva di lavoro ruota
diametro_t_p = 2*pcP[7]                             # [mm] diametro di testa pignone
diametro_t_r = 2*pcR[7]                             # [mm] diametro di testa ruota
diametro_pp_p = 2*pcP[8]                            # [mm] diametro di piede pignone
diametro_pp_r = 2*pcR[8]                            # [mm] diametro di piede ruota
                
b = pcP[11]                                         # [mm] larghezza di fascia  
eps_alpha = pcP[16]                                 # fattore di ricoprimento d'evolvente
eps_gamma = eps_alpha                               # fattore di ricoprimento totale
alpha_t = pcP[19]                                   # angolo di pressione trasversale di riferimento UNI 8862 tab4
alpha_wt = pcP[20]                                  # angolo di pressione trasversale di funzionamento UNI8862 tab4

# [mm] altezza dente pignone e ruota
h_p = pcP[5] + pcP[4]
h_r = pcR[5] + pcR[4]     

F_t = pcP[17]                                 # [N] forza tangenziale                         
F_r = pcP[18]                                 # [N] forza radiale
conn.close()




v = vel_rot_s*raggio_corretto_ruota               # [m/s] velocità lineare al diametro primitivo/di lavoro







################################################            ISO 6336-1: Fattori di influenza          ################################################  
################################################            5. Application factor  KA           ################################################ 
KA = 1.25                                                    # tab 1.1 vullo

specific_load = (F_t*KA)/b

st1 = f"CARICO SPECIFICO [N/mm] {int(specific_load)}"
out_file.write(st1 + '\n'), print(st1)





################################################            6. Internal dynamic factor  KV Method B (p20, 28)           ################################################  
####    Determination of resonance
m_red = rfi.massa_rel(diametro_t_p, diametro_l_p, diametro_pp_p ,density, u_real )              

st2 = f"MASSA RIDOTTA [kg/mm] {round(m_red, 5)}"
out_file.write(st2 + '\n'), print(st2)

####   9 Tooth stiffness parameters c′ and cγ p 70 (78)
qp, c_th = rfi.calcola_cthp(DentiP, DentiR, x_p, x_r)

CM = 0.8                                                    # eq 83

CR = 1                                                      # fig 19

CB_p = rfi.calcola_CB(ded_p, m_corr, theta_l)
CB_r = rfi.calcola_CB(ded_r, m_corr, theta_l)

CB = 0.5*(CB_p + CB_r)                                      # eq 87

c_prime = rfi.calcola_c_prime(c_th, CM, CR, CB, beta, specific_load)
c_gammaalpha = rfi.calcola_c_gammaalpha(c_prime, eps_alpha)                                               

tab_c_prime = [ ["Minimum value for the flexibility of a pair of teeth qp", round(qp, 3)],
        ["Theoretical single stiffness c'_th", round(c_th, 3)],
        ["Correction factor, CM", round(CM,2)],
         ["Gear blank factor, CR", round(CR,2)],
         ["Basic rack factor, CB", round(CB,2)], 
         ["Single stiffness, c'", round(c_prime,2)],
         ["Mesh stiffness, c_gammaalpha", round(c_gammaalpha,2)],]

st3 = "Stiffness parameters, c' & c_gamma"
st4 = tabulate(tab_c_prime)
out_file.write(st3 + '\n'), print(st3)
out_file.write(st4 + '\n'), print(st4)

####    Resonance
resonance_running_speed = rfi.calcola_vel_ris(DentiP, c_gammaalpha, m_red)                  

N = vel_rot_rpm/resonance_running_speed                     # resonance ratio eq 9 

####     Lower limit of resonance ratio     
N_s = 0.5 + 0.35*np.sqrt(specific_load/100)                 # eq. 11                                                                           

####     Individuazione zona di risonanza
st5 = f"Subcritical range {round(N, 3)}, {round(N_s, 3)}"
st6 = f"Main resonance range {round(N, 3)}, {round(N_s, 3)}"
st7 = f"Intermediate range {round(N, 3)}, {round(N_s, 3)}"
st8 = f"Supercritical range {round(N, 3)}, {round(N_s, 3)}"
if N <N_s:
    out_file.write(st5 + '\n'), print(st5)
if N>N_s and N<1.15:
    out_file.write(st6 + '\n'), print(st6)
if N>1.15 and N<1.5:
    out_file.write(st7 + '\n'), print(st7)
if N>1.5:
    out_file.write(st8 + '\n'), print(st8)

####    Dynamic factor in subcritical range                 
####    Coefficienti Bp Bf Bk
BP = rfi.calcola_BP(c_prime, f_bp, sigma_hlim, specific_load)
                            
BF = rfi.calcola_BF(c_prime, f_bp, sigma_hlim, specific_load)

BK = rfi.calcola_BK(c_prime, sigma_hlim, specific_load)

K,KV = rfi.calcola_KV(BP, BF, BK, N) 

tab_KV = [ ["Cv1", 0.32],
        ["Cv2", 0.34],
        ["Cv3", 0.23],
         ["BP", round(BP,2)],
         ["BF", round(BF,2)], 
         ["BK", round(BK,2)],
         ["K", round(K,2)],
         ["KV", round(KV,2)],]

st9 = "Dynamic factor in subcritical range"
st10 = tabulate(tab_KV)
out_file.write(st9 + '\n'), print(st9)
out_file.write(st10 + '\n'), print(st10)






################################################            7. Face load factors KHβ and KFβ   (p39, 47)         ################################################  
####    Determination of face load factor using Method C: KHβ-C         p 45 (53)  eq.39
F_m = F_t*KA*KV                                             # mean transverse tangential load

b_cal = (2* F_m * b)/F_t 

st11 = f"Per il calcolo usa il punto a): {round(b_cal/b,2)}"
st12 = f"Per il calcolo usa il punto b): {round(b_cal/b,2)}"
if  b_cal/b < 1:
    out_file.write(st11 + '\n'), print(st11)
else: 
    out_file.write(st12 + '\n'), print(st12)

KH_beta = rfi.calcola_KHbeta(sigma_hlim, F_beta_x, c_gammaalpha, F_m, b)

####    Determination of face load factor for tooth root stress using Method B or C: KFβ    
NF = 1/(1 + h_p/25 + (h_p/25)**2)                           # eq 69

KF_beta = (KH_beta)**(NF)                                   # eq 70 

tab_KHKF_beta = [["Mean transverse tangential load Fm", round(F_m,2)],
                 ["KH_beta", round(KH_beta,2)],
                 ["NF", round(NF,2)], 
                 ["KF_beta", round(KF_beta,2)]]

st13 = "Face load factors"
st14 = tabulate(tab_KHKF_beta)
out_file.write(st13 + '\n'), print(st13)
out_file.write(st14 + '\n'), print(st14)





################################################            8. Transverse load factors KHα and KFα    p(63, 71)        ################################################  
####    Determination of transverse load factor by calculation eq.71  p 64 (72) KHα = KFα
F_tH = F_t*KA*KV*KH_beta                                    # Tangential load in a transverse plane                                    

KH_alpha = rfi.calcola_Kalpha(sigma_hlim, f_bp, eps_gamma, c_gammaalpha, F_tH, b)
KF_alpha = KH_alpha 

tab_K_alpha = [["determinant tangential load in a transverse plane FtH", F_tH],
                 ["KH_alpha", round(KH_alpha,2)], 
                 ["KF_alpha", round(KF_alpha,2)]]

st15 = "Transverse load factors"
st16 = tabulate(tab_K_alpha)
out_file.write(st15 + '\n'), print(st15)
out_file.write(st16 + '\n'), print(st16)









################################################            ISO 6336-2: fattori di pitting          ################################################  
################################################            5 Zone factor, ZH        ################################################  
ZH = rfp.calcola_ZH(beta, alpha_t, alpha_wt)

tab_ZH = [["Angoli di pressione trasversale di riferimento [rad]", round(alpha_t,5)],
                 ["Angolo di pressione trasversale di funzionamento", round(alpha_wt,5)], 
                 ["Fattore di zona ZH", round(ZH,2)]]

st17 = "Zone factor"
st18 = tabulate(tab_ZH)
out_file.write(st17 + '\n'), print(st17)
out_file.write(st18 + '\n'), print(st18)




################################################            6 Single pair tooth contact factors, ZB and ZD        ################################################  
ZD = 1

# PB = (np.pi*diametro_pp_r)/DentiR

# eps_alpha = rfp.calcola_eps_alpha(diametro_t_p, diametro_t_r, diametro_pp_p, diametro_pp_r, theta_l, PB)

M1 = rfp.calcola_M1(diametro_t_p, diametro_t_r, diametro_pp_p, diametro_pp_r, eps_alpha, DentiP, DentiR, alpha_wt)

if M1 > 1:
    ZB = M1
else: 
    ZB = 1

tab_ZBZD = [["Fattore ZD", ZD],
                 ["Rapporto di condotta trasversale", round(eps_alpha,2)], 
                 ["Fattore M1", round(M1,2)], 
                 ["Fattore ZB", round(ZB,2)]]

st19 = "Single pair tooth contact factors, ZB and ZD"
st20 = tabulate(tab_ZBZD)
out_file.write(st19 + '\n'), print(st19)
out_file.write(st20 + '\n'), print(st20)




################################################            7 Elasticity factor, ZE      ################################################  
ZE = np.sqrt(0.175*E_modulus) #[sqrt(N/mm2)]

st21 = f"Elasticity factor, ZE: {round(ZE,2)}"
out_file.write(st21 + '\n'), print(st21)                # eq 21





################################################            8 Contact ratio factor, Zε      ################################################  
Z_eps = np.sqrt((4-eps_alpha)/3)

st22 = f"Contact ratio factor: {round(Z_eps,2)}"
out_file.write(st22 + '\n'), print(st22)              # eq 24





################################################            9 Helix angle factor, Zβ      ################################################ 
Z_beta = 1





################################################            11.2 Life factor ZNT (tab 2)     ################################################ 
ZNT_static = 1.6

ZNT_1010 = 0.85





################################################            12.3.1.1 Lubricant factor, ZL     ################################################ 
if sigma_hlim >1200:
    C_ZL = 0.91
elif sigma_hlim>850 and sigma_hlim <1200:
    C_ZL = (sigma_hlim)/4375 + 0.6357
elif sigma_hlim<850:
    C_ZL = 0.83

ZL = rfp.calcola_ZL(C_ZL, v_40)

st23 = f"Auxiliary factor CZL {round(C_ZL,2)}"
st24 = f"Lubricant factor, ZL: {round(ZL,2)}"
out_file.write(st23 + '\n'), print(st23)
out_file.write(st24 + '\n'), print(st24)

ZL_static = 1




################################################            12.3.1.2 Velocity factor, Zv     ################################################ 
ZV = rfp.calcola_ZV(C_ZL, v)

st25 = f"Velocity factor, Zv: {round(ZV,2)}"
out_file.write(st25 + '\n'), print(st25)

ZV_static = 1



################################################            12.3.1.3 Roughness factor, ZR     ################################################ 
rho_red = rfp.curvatura_relativa(diametro_pp_p, diametro_pp_r, alpha_wt)
Rz_10 = Rz*(10/rho_red)**(1/3)
C_ZR = 0.08                                                 # eq 51
ZR = (3/Rz_10)**(C_ZR)                                      # eq 48                                      

tab_ZR = [["Curvatura relativa", round(rho_red,2)], 
          ["Rugosità relativa Rz", Rz],
          ["Rugosità Rz10", round(Rz_10,2)], 
          ["Roughness factor, ZR", round(ZR,2)]]

st26 = "Roughness factor, ZR:"
st27 = tabulate(tab_ZR)
out_file.write(st26 + '\n'), print(st26)
out_file.write(st27 + '\n'), print(st27)

ZR_static = 1





################################################            13 Work hardening factor, ZW     ################################################ 
####    13.2.1 Surface-hardened pinion with through-hardened gear
Rz_H = rfp.calcola_RzH(Rz, rho_red, v_40, v)

####    13.2.1.2 ZW for reference and long life stress, determination by calculation
ZW = rfp.calcola_ZW(HB, Rz_H)

tab_ZW = [["Fattore Rz_h", round(Rz_H,2)],
          ["Durezza Brinell", HB],
          ["Work hardening factor, ZW", round(ZW,2)]]

st28 = "Work hardening factor, ZW:"
st29 = tabulate(tab_ZW)
out_file.write(st28 + '\n'), print(st28)
out_file.write(st29 + '\n'), print(st29)

ZW_static = 1





################################################            14 Size factor, ZX     ################################################ 
####    In this part of ISO 6336, ZX is taken to be 1,0.
ZX = 1









################################################            ISO 6336-3 tooth bending strength factors          ################################################
################################################           6.2 Calculation of the form factor, YF: Method B          ################################################
#### VULLO P 172 The determination of the value of YF is based on the nominal tooth shape

####    6.2.1 Tooth root normal chord, sFn, radius of root fillet, ρF, bending moment arm, hFe
####    Tooth root normal chord, sFn
#### cARATTERISTICHE DELA DENTIERA UTENSILE 1.25/0.38/1.0 ISO 53.2A.
# protuberanza di sottotaglio
spr = 0                                                     # gears not undercut

####    Reinizializzo il modulo: ora sto facendo calcoli relativia alla realizzazione della ruota, unificata.
modulo = m_uni
print(modulo)

####    Raggio di curvatura alla base del dente: raggio di testa utensile / raggio di fondo dentiera di riferimento
rho_fp = 0.375*modulo

####    addendum utensile / dedendum dentiera di riferimento
u = 1.25*modulo

####    Calcolo dei fattori ausiliari
E_factor, G_factor, H_factor = rff.calcola_EGH(rho_fp, u, modulo, x_r, theta_p, spr, DentiR)

theta_aus, iterazioni = rff.calcola_theta_aus(0.001, np.pi/6, DentiR, H_factor, G_factor)

####    tooth normal cord: spessore critico
sfn, sfn_norm = rff.calcola_sFn(G_factor, rho_fp, modulo, theta_aus, DentiR)

####    radius of root fillet, ρF: raggio di curvatura della sez. critica
rho_F = rff.calcola_rhoF(modulo, DentiR, theta_aus, G_factor, rho_fp)

tab_YF1 = [["Raggio di curvatura alla base", rho_fp ],
          ["Fattore ausiliario E", round(E_factor, 4)],
          ["Fattore ausiliario G", round(G_factor,4)],
          ["Fattore ausiliario H", round(H_factor,4)],
          ["Fattore ausiliario theta", round(theta_aus,5)],
          ["Numero di iterazioni", round(iterazioni,2)],
          ["Tooth root normal chord, sFn", round(sfn,2)],
          ["Tooth root normal chord, normalized", round(sfn_norm,2)],
          ["Radius of root fillet rhoF", round(rho_F,2)]]

st30 = "Form factor, YF tab1/2:"
st31 = tabulate(tab_YF1)
out_file.write(st30 + '\n'), print(st30)
out_file.write(st31 + '\n'), print(st31)

####    Bending moment arm, hFe
####    Reinizializzazione variabili
####    pitch diameter in normal section
dn = modulo*DentiR                                           # eq 24 
####    Base diameter in normal section                                            
dbn = dn*np.cos(theta_p)                                    # eq 26
####  Tip diameter in normal section  
dan = diametro_t_r                      # eq ??
#### Form-factor pressure angle (pressure angle at the outer point of single pairtooth contact), 
#### Diameter of circle through outer point of single pair tooth contact in normal plane or outer pitch diameter in normal section
alpha_en, den = rff.calcola_alphaen_den(dan, dbn, dn, theta_p, beta, DentiR, eps_alpha)
#### Tooth thickness half angle
gamma_e = rff.calcola_gamma_e(x_r, DentiR, theta_p, alpha_en)
#### Load direction angle,
alpha_fen = rff.calcola_alphaFen(alpha_en, gamma_e)                              # eq 31
hFE, hFE_norm = rff.calcola_hFe((gamma_e), (alpha_fen), den, modulo, (theta_aus), DentiR, G_factor, rho_fp)


tab_YF2 = [["dn", round(dn,2) ],
          ["dbn", round(dbn,2)],
          ["dan", round(dan,2)],
          ["den", round(den,2)],
          ["alpha_en", round(alpha_en,5)],
          ["gamma_e", round(gamma_e,5)],
          ["alpha_fen", round(alpha_fen,5)],
          ["bending moment arm, hFe", round(hFE,2)],
          ["bending moment arm, normalized", round(hFE_norm,2)]]

st32 = "Form factor, YF tab2/2:"
st33 = tabulate(tab_YF2)
out_file.write(st32 + '\n'), print(st32)
out_file.write(st33 + '\n'), print(st33)        

YF = rff.calcola_YF(hFE_norm, sfn_norm, alpha_fen, theta_p)

st34 = f"Form factor, YF FINAL: {round(YF,2)}"
out_file.write(st34 + '\n'), print(st34)        





################################################           7.2 Stress correction factor, YS: Method B          ################################################

YS, L, qs = rff.calcola_YS(sfn, hFE, rho_F)

tab_YS = [["PARAMETRO D'INTAGLIO qs", round(qs,2) ],
          ["L", round(L,2)],
          ["FATTORE DI CORREZIONE DELLA TENSIONE YS", round(YS,2)],
          ]

st35 = "Stress correction factor, YS:"
st36 = tabulate(tab_YS)         
out_file.write(st35 + '\n'), print(st35) 
out_file.write(st36 + '\n'), print(st36) 

####    Tensione a fondo dente preliminare - normal tooth load
sigma_F1 = (F_t/(b*modulo))*YF*YS

st37 = f"Tensione a fondo dente preliminare [N/mm] = [MPa]: {round(sigma_F1,2)}"
out_file.write(st37 + '\n'), print(st37)





################################################           8 Helix angle factor, Yβ         ################################################
Y_beta = 1                                                  # DENTI DRITTI





################################################           9 Rim thickness factor, YB         ################################################
####    RICAVO AL CONTRARIO, AFFICHÈ YB = 1 DEVE ESSERE
sr = int(1.2*h_r) +1

st38 = f"MINIMO SPESSORE DELL'ANELLO [mm] {sr}"
out_file.write(st38 + '\n'), print(st38)

d_alberomax = diametro_pp_r-(2*sr)

st39 = f"MASSIMO DIAMETRO DELL'ALBERO [mm] {int(d_alberomax)}"
out_file.write(st39 + '\n'), print(st39)

YB = 1





################################################           10 Deep tooth factor, YDT         ################################################
eps_alpha_n = eps_alpha

st40 = f"Virtual conctat ratio of a spur gear {round(eps_alpha_n,2)}"
out_file.write(st40 + '\n'), print(st40)

YDT = 1





################################################           12 Life factor, YNT         ################################################
YNT = 1                                                     # fig 9 p20

YNT_static = 2.5                                            # tab1 p21





################################################           13 Sensitivity factor, YδT, and relative notch sensitivity factor, Yδ rel T         ################################################
Y_delta_relT = 0.98                                         # fig10 p24

Y_delta_relT_static = 0.72                                  # fig11 p25

Y_delta_t = 1.12                                            # fig12 p26

Y_delta_t_static = 1.25                                     # fig13 p27





################################################           14 Surface factors, YR, YRT, and relative surface factor, YR rel T         ################################################
####    I fattori YR, YRT, possono essere inclusi nel diagramma S-N pertanto, come
####    espresso dalla normativa vengono considerati unitari, fatta eccezione per y!"
####    14.3.2.1 YR rel T for reference stress
YR_rel_T = 1.07                                             # eq 57

####    14.2.2 Method B
YR_rel_T_static = 1





################################################           15 Size factor, YX         ################################################
YX = 1                                                      # tab 3 pag 32

YX_static = YX










################################################           VERIFICHE        ################################################
####    5.3.2 Permissible bending stress, σFP: Method B
sigma_FP_static = (sigma_flim/SF_lim)*YS*YNT_static*Y_delta_relT_static*YR_rel_T_static*YX_static
sigma_FP = (sigma_flim/SF_lim)*YS*YNT*Y_delta_relT*YR_rel_T*YX

####    5.4 Permissible contact stress, σHP
sigma_HP_static = (sigma_hlim/SH_lim)*ZNT_static*ZL_static*ZV_static*ZR_static*ZW_static*ZX
sigma_HP = (sigma_hlim/SH_lim)*ZNT_1010*ZL*ZV*ZR*ZW*ZX





################################################           VERIFICA A FLESSIONE        ################################################
####    Tooth root stress, σF
sigma_F0 = (F_t/(m_uni*b))*YF*YS*Y_beta*YB*YDT
sigma_F = sigma_F0*KA*KV*KF_beta*KF_alpha

####    STATICA
str1 = f"La verifica A FLESSIONE STATICA è superata! {round(sigma_F, 2)}, {round(sigma_FP_static, 2)}"
str2 = f"La verifica A FLESSIONE STATICA NON è superata! {round(sigma_F, 2)}, {round(sigma_FP_static, 2)}"
if sigma_F<sigma_FP_static:
    out_file.write(str1 + '\n'), print(str1)
else:
    out_file.write(str2 + '\n'), print(str2)

####    A FATICA
st41 = f"La verifica A FLESSIONE è superata! {round(sigma_F, 2)}, {round(sigma_FP, 2)}"
st42 = f"La verifica A FLESSIONE NON è superata! {round(sigma_F, 2)}, {round(sigma_FP, 2)}"
if sigma_F<sigma_FP:
    out_file.write(st41 + '\n'), print(st41)
else:
    out_file.write(st42 + '\n'), print(st42)





################################################           VERIFICA A PITTING        ################################################
####    nominal contact stress
sigma_H0 = ZH*ZE*Z_eps*Z_beta*np.sqrt(F_t/(diametro_p_p*b) * ((u_real+1)/u_real))
sigma_H1 = ZB*sigma_H0*np.sqrt(KA*KV*KH_beta*KH_alpha)
sigma_H2 = ZD*sigma_H0*np.sqrt(KA*KV*KH_beta*KH_alpha)

####    STATICA
str3 = f"La verifica A PITTING STATICA è superata! {round(sigma_H1, 2)}, {round(sigma_H1, 2)}, {round(sigma_HP_static, 2)}"
str4 = f"La verifica A PITTING STATICA NON è superata! {round(sigma_H1, 2)}, {round(sigma_H1, 2)}, {round(sigma_HP_static, 2)}"
if sigma_H1<sigma_HP_static and sigma_H2<sigma_HP_static:
    out_file.write(str3 + '\n'), print(str3)
else:
    out_file.write(str4 + '\n'), print(str4)

####    A FATICA    
st43 = f"La verifica A PITTING è superata! {round(sigma_H1, 2)}, {round(sigma_H1, 2)}, {round(sigma_HP, 2)}"
st44 = f"La verifica A PITTING NON è superata! {round(sigma_H1, 2)}, {round(sigma_H1, 2)}, {round(sigma_HP, 2)}"
if sigma_H1<sigma_HP and sigma_H2<sigma_HP:
    out_file.write(st43 + '\n'), print(st43)
else:
    out_file.write(st44 + '\n'), print(st44)









################################################            DATABASE        ################################################
out_file.close() 






####    Influence factors
influence_factors = {
    'fattori':["KA", 
               "KV", 
               "KHbeta",
                "KFbeta", 
                "KHalpha", 
                "KFalpha"],
    'valori':[ KA, 
              round(KV,2), 
              round(KH_beta,2),
              round(KF_beta,2), 
              round(KH_alpha,2),
              round(KF_alpha,2)]
}
index_labels1 = ['r1', 'r2', 'r3', 'r4', 'r5', 'r6']

IF = pd.DataFrame(influence_factors, index=index_labels1)

# Connessione al database SQLite
conn = sqlite3.connect("influence_factorsR.db")

# Salva il DataFrame nel database
IF.to_sql('influence_factorsR', conn, if_exists='replace', index_label='index')

# Chiudi la connessione
conn.close()

# Tabella su excel sempre facilmente consultabile
IF.to_excel('influence_factorsR.xlsx', index=True)




####    pitting factors
pitting_factors = {
    'fattori':["Zone factor ZH", 
               "Single pair tooth contact factor ZB", 
               "Single pair tooth contact factor ZD",
                "Elasticity factor ZE", 
                "Contact ratio factor Zeps", 
                "Helix angle factor Zbeta", 
                "Life factor ZNT(10^10)",
                "Life factor ZNT(static)", 
                "Lubricant factor ZL",
                "Lubricant factor ZL(static)", 
                "Velocity factor ZV", 
                "Velocity factor ZV(static)", 
                "Roughness factor ZR", 
                "Roughness factor ZR(static)", 
                "Work hardening factor ZW", 
                "Work hardening factor ZW(static)",
                "Size factor ZX"],

    'valori':[round(ZH,2), 
              round(ZB,2), 
              round(ZD,2),
              round(ZE,2), 
              round(Z_eps,2),
              round(Z_beta,2), 
              round(ZNT_1010,2),
              ZNT_static,
              round(ZL,2),
              ZL_static, 
              round(ZV,2),
              ZV_static,
              round(ZR,2),
              ZR_static,
              round(ZW,2), 
              ZW_static, 
              round(ZX, 2)]
               
}
index_labels2 = ['r1', 'r2', 'r3', 'r4', 'r5', 
                 'r6', 'r7', 'r8', 'r9', 'r10', 
                 'r11', 'r12', 'r13', 'r14', 'r15', 
                 'r16', 'r17']

PF = pd.DataFrame(pitting_factors, index=index_labels2)

# Connessione al database SQLite
conn = sqlite3.connect("pitting_factorsR.db")

# Salva il DataFrame nel database
PF.to_sql('pitting_factorsR', conn, if_exists='replace', index_label='index')

# Chiudi la connessione
conn.close()

# Tabella su excel sempre facilmente consultabile
PF.to_excel('pitting_factorsR.xlsx', index=True)




####    bending factors
bending_factors = {
    'fattori':["Form factor YF", 
               "Stress correction factor YS", 
               "Helix angle factor Ybeta",
                "Rim thickness factor YB", 
                "Deep tooth factor YDT", 
                "Life factor YNT", 
                "Life factor YNT(static)", 
                "Sensitivity factor YdeltaT",
                "Sensitivity factor YdeltaT(static)", 
                "Relative notch sensitivity factorYdeltarelT", 
                "Relative notch sensitivity factorYdeltarelT(static)", 
                "Relative surface factor YRrelT", 
                "Relative surface factor YRrelT(static)", 
                "Size factor YX", 
                "Size factor YX(static)"],

    'valori':[round(YF,2), 
              round(YS,2), 
              round(Y_beta,2),
              round(YB,2), 
              round(YDT,2),
              round(YNT,2), 
              YNT_static,
              Y_delta_t,
              Y_delta_t_static,
              Y_delta_relT, 
              Y_delta_relT_static,
              YR_rel_T,
              YR_rel_T_static,
              YX,
              YX_static]
               
}
index_labels3 = ['r1', 'r2', 'r3', 'r4', 'r5', 
                 'r6', 'r7', 'r8', 'r9', 'r10', 
                 'r11', 'r12', 'r13', 'r14', 'r15']

BF = pd.DataFrame(bending_factors, index=index_labels3)

# Connessione al database SQLite
conn = sqlite3.connect("bending_factorsR.db")

# Salva il DataFrame nel database
BF.to_sql('bending_factorsR', conn, if_exists='replace', index_label='index')

# Chiudi la connessione
conn.close()

# Tabella su excel sempre facilmente consultabile
BF.to_excel('bending_factorsR.xlsx', index=True)









################################################            ORGANIZZAZIONE FILE        ################################################
# Ottieni il percorso della directory in cui si trova lo script
percorso_script = os.path.dirname(os.path.abspath(__file__))

# Costruisci il percorso completo della cartella di origine
cartella_origine = os.path.join(percorso_script)

# Crea una cartella per le figure, tabelle e log all'interno della directory dello script (senza errori se esistono già)
os.makedirs(os.path.join(cartella_origine, 'figure'), exist_ok=True)
os.makedirs(os.path.join(cartella_origine, 'tabelle'), exist_ok=True)
os.makedirs(os.path.join(cartella_origine, 'log'), exist_ok=True)
os.makedirs(os.path.join(cartella_origine, 'database'), exist_ok=True)


# Elenca tutti i file nella cartella di origine
files = os.listdir(cartella_origine)

# Sposta i file nelle cartelle corrispondenti
for file in files:
    if file.endswith('.png') or file.endswith('.jpg'):
        shutil.move(os.path.join(cartella_origine, file), os.path.join(cartella_origine, 'figure', file))
    elif file.endswith('.xlsx'):
        shutil.move(os.path.join(cartella_origine, file), os.path.join(cartella_origine, 'tabelle', file))
    elif file.endswith('.txt'):
        shutil.move(os.path.join(cartella_origine, file), os.path.join(cartella_origine, 'log', file))
    elif file.endswith('.db'):
        shutil.move(os.path.join(cartella_origine, file), os.path.join(cartella_origine, 'database', file))


end_time = datetime.now()
print('Elapsed Time: {}'.format(end_time - start_time))













